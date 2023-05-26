<<<<<<< HEAD
=======
## pyinstaller --noconsole serials.py --onefile --hidden-import openpyxl.cell._writer

>>>>>>> 8092570415752499015f52e25b44fe7c4dfaffe3
## pyinstaller --noconsole --noconfirm serials.py --onefile --hidden-import openpyxl.cell._writer

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import difflib
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pyhtml2pdf import converter
import os

file_paths = []
target_columns = ['Unit', 'Fridge', 'Range', 'Microwave', 'Dishwasher', 'Washer', 'Dryer']
tab_names = ["Preview", "Details"]
data_loaded_label = None


def close_match(col_name, column_list):
    best_match = difflib.get_close_matches(col_name, column_list, n=1, cutoff=0.6)
    return best_match[0] if best_match else col_name

def process_data():
    combined_df = pd.DataFrame()

    for file_path in file_paths:
        df = pd.read_excel(file_path)
        df.columns = [col.lower() for col in df.columns]
        df.columns = [close_match(col, target_columns) for col in df.columns]

        for col in target_columns:
            if col in df.columns:
                df[col] = df[col].astype(str).str.upper().str.replace('-', '').str.replace(',', '').str.replace('.', '').str.replace(' ', '').str.replace('AND', 'N').str.replace('IS', '').str.replace('ARE', 'R').str.replace('IF', 'F').str.replace('SEA', 'C').str.replace('FP', 'FB').str.replace('#', '').str.replace(':', '')

        combined_df = pd.concat([combined_df, df])

    combined_df = combined_df.sort_values(by='Unit', ascending=True)
    combined_df = combined_df.reset_index(drop=True)

    return combined_df

def save_data():
    combined_df = process_data()

    if not combined_df.empty:
        combined_df.drop_duplicates(inplace=True)
        output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="Processed Serial Numbers")
        if output_file:
            combined_df.replace({pd.NA: 'N/A', 'NAN': 'N/A'}, inplace=True)
            combined_df.to_excel(output_file, index=False)
            auto_fit_columns(output_file)

            html_file = output_file.replace(".xlsx", ".html")
            combined_df.to_html(html_file, index=False)

            # Add CSS styling
            css = """
            <style>
            table {
                border-collapse: collapse;
                width: 100%;
            }
            
            th, td {
                border: 1px solid black;
                padding: 8px;
                text-align: center;
            }
            </style>
            """
            
            # Read the HTML file
            with open(html_file, 'r') as file:
                html_content = file.read()

            # Insert the CSS styling
            html_content_with_css = css + html_content

            # Write the modified HTML file
            with open(html_file, 'w') as file:
                file.write(html_content_with_css)

            output_pdf_file = filedialog.asksaveasfilename(defaultextension=".pdf", initialfile="Serials")
            if output_pdf_file:
                converter.convert(html_file, output_pdf_file)
                messagebox.showinfo("Success", "Processing completed successfully!")

                root.destroy()
                os.remove(html_file)
    else:
        messagebox.showwarning("Warning", "No files selected.") 


def pick_files():
    global data_loaded_label

    files = filedialog.askopenfilenames()
    if files:
        file_paths.clear()
        file_paths.extend(files)
        data_loaded_label1 = tk.Label(root, text="DATA LOADED. You can now save the files.", fg="red")
        data_loaded_label1.pack(pady=10)

        data_loaded_label2 = tk.Label(root, text="Program will close after completed", fg="blue")
        data_loaded_label2.pack(pady=10)


def process_files():
    if file_paths:
        save_data()
    else:
        messagebox.showwarning("Warning", "No files selected.") 


def auto_fit_columns(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        col_letter = get_column_letter(column[0].column)
        ws.column_dimensions[col_letter].width = adjusted_width
    wb.save(file_path)
    wb.close()

root = tk.Tk()
root.title("Serial Number Processing")

root.geometry("350x250")
root.configure(background='darkgray')

frame = tk.Frame(root, bg='darkgray')
frame.place(relx=0.5, rely=0.5, anchor='center')

button_pick = tk.Button(root, text="Pick Excel File(s)", command=pick_files)
button_pick.pack(pady=20)

button_process = tk.Button(root, text="Process and Save", command=process_files)
button_process.pack(pady=10)

data_loaded_label = tk.Label(root) # Initialize the data_loaded_label

<<<<<<< HEAD
root.mainloop()
=======
root.mainloop()
>>>>>>> 8092570415752499015f52e25b44fe7c4dfaffe3
