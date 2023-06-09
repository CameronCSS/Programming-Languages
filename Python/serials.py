## pyinstaller --windowed serials.py --onefile --hidden-import openpyxl.cell._writer --hidden-import pyhtml2pdf --hidden-import pdfkit

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import difflib
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pyhtml2pdf import converter
import os
import threading

file_paths = []
target_columns = ['Unit', 'Fridge', 'Range', 'Microwave', 'Dishwasher', 'Washer', 'Dryer']
tab_names = ["Preview", "Details"]
data_loaded_label1 = None
data_loaded_label2 = None
processing_label = None

def close_match(col_name, column_list):
    best_match = difflib.get_close_matches(col_name, column_list, n=1, cutoff=0.6)
    return best_match[0] if best_match else col_name

def process_data():
    combined_df = pd.DataFrame()

    for file_path in file_paths:
        try:
            df = pd.read_excel(file_path)
        except FileNotFoundError:
            messagebox.showerror("Error", f"File not found: {file_path}")
            continue  # Skip to the next file
        except PermissionError:
            messagebox.showerror("Error", f"No permission to read the file: {file_path}")
            continue  # Skip to the next file
        except IOError:
            messagebox.showerror("Error", f"IOError occurred while reading the file: {file_path}")
            continue  # Skip to the next file
        except OSError:
            messagebox.showerror("Error", f"OSError occurred while reading the file: {file_path}")
            continue  # Skip to the next file
        except Exception as e:
            messagebox.showerror("Error", f"An unknown error occurred while reading the file: {file_path}. Error: {str(e)}")
            continue  # Skip to the next file
        df.columns = [col.lower() for col in df.columns]
        df.columns = [close_match(col, target_columns) for col in df.columns]

        for col in target_columns:
            if col in df.columns:
                df[col] = df[col].astype(str).str.upper().str.replace('-', '').str.replace(',', '').str.replace('.', '').str.replace(' ', '').str.replace('AND', 'N').str.replace('IS', '').str.replace('ARE', 'R').str.replace('IF', 'F').str.replace('SEA', 'C').str.replace('FP', 'FB').str.replace('#', '').str.replace(':', '')

        combined_df = pd.concat([combined_df, df])

    combined_df = combined_df.sort_values(by='Unit', ascending=True)
    combined_df = combined_df.reset_index(drop=True)

    return combined_df

def save_data():
    processing_completed = False  # Flag to track if processing completed successfully
    try:
        combined_df = process_data()
        global data_loaded_label1
        global data_loaded_label2
        global processing_label

        # Hide the old labels
        if data_loaded_label1:
            data_loaded_label1.pack_forget()
        if data_loaded_label2:
            data_loaded_label2.pack_forget()

        if not combined_df.empty:
            combined_df.drop_duplicates(inplace=True)
            
            # Ask for output file name for the PDF and use it for temporary Excel and HTML files
            output_pdf_file = filedialog.asksaveasfilename(defaultextension=".pdf", initialfile="Serials")
            if not output_pdf_file:
                output_pdf_file = "processed_serials.pdf"

            if output_pdf_file:
                processing_label = tk.Label(root, text="Processing... Please Wait", fg="blue")
                processing_label.pack(pady=10)
                root.update()  # Force the window to update

                output_file = output_pdf_file.replace(".pdf", ".xlsx")
                html_file = output_pdf_file.replace(".pdf", ".html")

                combined_df.replace({pd.NA: 'N/A', 'NAN': 'N/A'}, inplace=True)
                combined_df.to_excel(output_file, index=False)
                auto_fit_columns(output_file)

                combined_df.to_html(html_file, index=False)

                # Add CSS styling
                css = """
                <style>
                table {
                    border-collapse: collapse;
                    width: 100%;
                }
                
                th, td {
                    border: 1px solid black;
                    padding: 8px;
                    text-align: center;
                }
                </style>
                """
                
                # Read the HTML file
                with open(html_file, 'r') as file:
                    html_content = file.read()

                # Insert the CSS styling
                html_content_with_css = css + html_content

                try:
                # Write the modified HTML file
                    with open(html_file, 'w') as file:
                        file.write(html_content_with_css)
                except IOError:
                    messagebox.showerror("Error", f"IOError occurred while trying to write to the file: {html_file}")
                except OSError:
                    messagebox.showerror("Error", f"OSError occurred while trying to write to the file: {html_file}")
                except Exception as e:
                    messagebox.showerror("Error", f"An unknown error occurred while trying to write to the file: {html_file}. Error: {str(e)}")

                converter.convert(html_file, output_pdf_file)
                messagebox.showinfo("Success", "Processing completed successfully!")

                processing_completed = True  # Processing completed successfully

                root.after(1000, root.destroy)  # Delay closing the app by 1 second

                # Attempt to delete the temporary files
                def delete_files():
                    try:
                        if os.path.exists(html_file):
                            os.remove(html_file)
                        if os.path.exists(output_file):
                            os.remove(output_file) 
                    except Exception as e:
                        print(f"Failed to delete files: {e}")

                root.after(500, delete_files)  # Try to delete the files after 0.5 seconds

        else:
            messagebox.showwarning("Warning", "No files selected.")
    except Exception as e:
        # An error occurred, show an error message
        messagebox.showerror("Error", f"An error occurred while processing the files: {str(e)}")
    finally:
        if not processing_completed:
            # Delete the temporary files if processing did not complete successfully
            try:
                if os.path.exists(html_file):
                    os.remove(html_file)
                if os.path.exists(output_file):
                    os.remove(output_file) 
            except Exception as e:
                print(f"Failed to delete files: {e}")
            root.quit()



def pick_files():
    global data_loaded_label1
    global data_loaded_label2

    files = filedialog.askopenfilenames(filetypes=(("Excel Files", "*.xls *.xlsx *.xlsm"),("All Files", "*.*")))

    # check if selected files are Excel files
    for file in files:
        if not file.lower().endswith(('.xls', '.xlsx', '.xlsm')):
            messagebox.showwarning("Invalid file", f"Selected file {file} is not an Excel file.")
            return  # Do not proceed with invalid files

    if files:
        file_paths.clear()
        file_paths.extend(files)

        # Destroy the old labels if they exist
        if data_loaded_label1:
            data_loaded_label1.destroy()
        if data_loaded_label2:
            data_loaded_label2.destroy()

        # Create new labels
        data_loaded_label1 = tk.Label(root, text="DATA LOADED", fg="red")
        data_loaded_label1.pack(pady=10)
        
        data_loaded_label2 = tk.Label(root, text="You can now Save.", fg="blue")
        data_loaded_label2.pack(pady=10)


def process_files():
    if file_paths:
        threading.Thread(target=save_data).start()
    else:
        messagebox.showwarning("Warning", "No files selected.") 


def auto_fit_columns(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        col_letter = get_column_letter(column[0].column)
        ws.column_dimensions[col_letter].width = adjusted_width
    wb.save(file_path)
    wb.close()

root = tk.Tk()
root.title("Serial Number Processing")

root.geometry("350x250")
root.configure(background='darkgray')

frame = tk.Frame(root, bg='darkgray')
frame.place(relx=0.5, rely=0.5, anchor='center')

button_pick = tk.Button(root, text="Pick Excel File(s)", command=pick_files)
button_pick.pack(pady=20)

button_process = tk.Button(root, text="Process and Save", command=process_files)
button_process.pack(pady=10)

root.mainloop()
